from datetime import datetime
from zoneinfo import ZoneInfo
from pathlib import Path
import os

import requests
from openpyxl import Workbook, load_workbook
import yfinance as yf


EXCEL_BESTAND = Path("prijzen.xlsx")
TIJDZONE = ZoneInfo("Europe/Amsterdam")


# Alles in EUR opslaan
TICKERS_EUR = {
    "ASML EUR": "ASML.AS",
    "Pharming EUR": "PHARM.AS",
    "TDIV EUR": "TDIV.AS",
    "EUNL EUR": "EUNL.AS",
    "VUAA EUR": "VUAA.AS",
    "Magnum EUR": "7RM.DU",
    "DFNS EUR": "DFNS.PA",
}


# USD-assets die we naar EUR omrekenen
TICKERS_USD_TO_EUR = {
    "AGNC EUR": "AGNC",
    "NVIDIA EUR": "NVDA",
}


def haal_crypto_prijzen_op():
    url = "https://api.coingecko.com/api/v3/simple/price"
    params = {
        "ids": "bitcoin,ethereum",
        "vs_currencies": "eur"
    }
    response = requests.get(url, params=params, timeout=30)
    response.raise_for_status()
    data = response.json()

    return {
        "Bitcoin EUR": round(float(data["bitcoin"]["eur"]), 2),
        "Ethereum EUR": round(float(data["ethereum"]["eur"]), 2),
    }


def haal_eurusd_op():
    fx = yf.Ticker("EURUSD=X")
    hist = fx.history(period="5d", interval="1d")

    if hist.empty:
        raise ValueError("Geen EUR/USD data gevonden")

    return float(hist["Close"].dropna().iloc[-1])


def haal_metalen_op(eurusd):
    goud_hist = yf.Ticker("GC=F").history(period="5d", interval="1d")
    zilver_hist = yf.Ticker("SI=F").history(period="5d", interval="1d")
    platina_hist = yf.Ticker("PL=F").history(period="5d", interval="1d")

    if goud_hist.empty or zilver_hist.empty or platina_hist.empty:
        raise ValueError("Geen metaaldata gevonden")

    goud_usd_oz = float(goud_hist["Close"].dropna().iloc[-1])
    zilver_usd_oz = float(zilver_hist["Close"].dropna().iloc[-1])
    platina_usd_oz = float(platina_hist["Close"].dropna().iloc[-1])

    ounce_naar_kg = 32.1507466

    return {
        "Goud EUR/kg": round((goud_usd_oz * ounce_naar_kg) / eurusd, 2),
        "Zilver EUR/kg": round((zilver_usd_oz * ounce_naar_kg) / eurusd, 2),
        "Platina EUR/kg": round((platina_usd_oz * ounce_naar_kg) / eurusd, 2),
    }


def haal_aandelen_en_etfs_op(eurusd):
    resultaten = {}

    for naam, ticker in TICKERS_EUR.items():
        hist = yf.Ticker(ticker).history(period="5d", interval="1d")

        if hist.empty:
            print(f"Geen data voor {naam} ({ticker})")
            resultaten[naam] = None
            continue

        prijs = float(hist["Close"].dropna().iloc[-1])
        resultaten[naam] = round(prijs, 2)

    for naam, ticker in TICKERS_USD_TO_EUR.items():
        hist = yf.Ticker(ticker).history(period="5d", interval="1d")

        if hist.empty:
            print(f"Geen data voor {naam} ({ticker})")
            resultaten[naam] = None
            continue

        prijs_usd = float(hist["Close"].dropna().iloc[-1])
        prijs_eur = prijs_usd / eurusd
        resultaten[naam] = round(prijs_eur, 2)

    return resultaten


def maak_excel_als_nodig(kolommen):
    if not EXCEL_BESTAND.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Prijzen"
        ws.append(["Datum", "Tijd"] + kolommen)
        wb.save(EXCEL_BESTAND)


def laatste_datum(ws):
    if ws.max_row <= 1:
        return None
    return ws.cell(row=ws.max_row, column=1).value


def main():
    nu = datetime.now(TIJDZONE)
    datum = nu.strftime("%Y-%m-%d")
    tijd = nu.strftime("%H:%M:%S")

    github_event_name = os.getenv("GITHUB_EVENT_NAME", "")

    if github_event_name != "workflow_dispatch" and nu.hour != 0:
        print(f"Niet uitgevoerd: lokale NL-tijd is {tijd}")
        return

    eurusd = haal_eurusd_op()

    crypto = haal_crypto_prijzen_op()
    aandelen_etfs = haal_aandelen_en_etfs_op(eurusd)
    metalen = haal_metalen_op(eurusd)

    alle_data = {}
    alle_data.update(crypto)
    alle_data.update(aandelen_etfs)
    alle_data.update(metalen)

    kolommen = list(alle_data.keys())

    maak_excel_als_nodig(kolommen)

    wb = load_workbook(EXCEL_BESTAND)
    ws = wb["Prijzen"]

    if laatste_datum(ws) == datum:
        print(f"Bestand was al bijgewerkt voor {datum}")
        return

    ws.append([datum, tijd] + [alle_data[k] for k in kolommen])
    wb.save(EXCEL_BESTAND)

    print(f"Toegevoegd voor {datum} {tijd}")


if __name__ == "__main__":
    main()
